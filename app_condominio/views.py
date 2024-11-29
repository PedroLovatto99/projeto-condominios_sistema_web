from django.shortcuts import render
from .models import *
from django.views.generic import *
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import redirect
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
import datetime
from babel.dates import format_datetime, format_date


@method_decorator(login_required(login_url='login'), name='dispatch')
class BlocosListView(ListView):
    model = Bloco
    template_name = "blocos-lista.html"
    context_object_name = "blocos"
    ordering = ['numero']

@method_decorator(login_required(login_url='login'), name='dispatch')
class BlocoDetail(DetailView):
    model = Bloco
    template_name = 'bloco.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        apartamentos = Apartamento.objects.filter(bloco=self.object).order_by("numero")
        residentes = Residente.objects.filter(apartamento__in=apartamentos)
        
        for residente in residentes:
            residente.total_boleto = (residente.valor_aluguel or 0) + \
                                     (residente.valor_condominio or 0) + \
                                     (residente.valor_gas or 0) + \
                                     (residente.outros or 0)

        context['apartamentos'] = apartamentos
        context['residentes'] = residentes

        return context

def login_view(request):
    login_form = AuthenticationForm(data=request.POST or None)
    if request.method == 'POST':
        if login_form.is_valid():
            username = login_form.cleaned_data.get('username')
            password = login_form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('blocos_lista') 
    return render(request, 'login.html', {'login_form': login_form})

def logout_view(request):
    logout(request)
    return redirect('login')


def formatar_data(data):
    return format_date(data, format="short", locale="pt_BR") if data else ""

def gerar_excel(request, bloco_id):
    # Obter os apartamentos e residentes relacionados ao bloco
    apartamentos = Apartamento.objects.filter(bloco_id=bloco_id)
    residentes = Residente.objects.filter(apartamento__in=apartamentos)

    # Preparar os dados para o DataFrame
    data = []
    for apartamento in apartamentos:
        for residente in residentes:
            if residente.apartamento == apartamento:
                total_boleto = (residente.valor_aluguel or 0) + (residente.valor_condominio or 0) + (residente.valor_gas or 0) + (residente.outros or 0)

                entry = {
                    'Apartamento': apartamento.numero,
                    'Residente': residente.nome,
                    'Email': residente.email,
                    'Telefone': residente.telefone,
                    'Data Início Contrato': formatar_data(residente.data_inicio),
                    'Data Fim Contrato': formatar_data(residente.data_fim),
                    'Prorrogação': residente.prorrogacao,
                    'Número Cadastro': residente.numero_cadastro,
                    'Valor Aluguel': residente.valor_aluguel,
                    'Valor Condomínio': residente.valor_condominio,
                    'Outros Valores': residente.outros,
                    'Valor Gás': residente.valor_gas,
                    'Data Vencimento Boleto': formatar_data(residente.data_vencimento),
                    'Unidade Consumidora': residente.unidade_consumidora,
                    'Total Boleto': total_boleto,
                }
                if request.user.is_superuser:
                    entry['CPF/CNPJ'] = residente.cpf_cnpj
                data.append(entry)

    # Criar o DataFrame com os dados
    df = pd.DataFrame(data)

    # Criar um buffer de memória
    buffer = BytesIO()

    # Gerar o arquivo Excel no buffer
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Bloco')

        # Obter o workbook e a aba
        workbook = writer.book
        worksheet = writer.sheets['Bloco']

        # Adicionar título com o número do bloco
        bloco = Bloco.objects.get(pk=bloco_id)
        worksheet.insert_rows(1)  # Inserir uma nova linha no topo
        worksheet['A1'] = f'Bloco {bloco.numero}'  # Adicionar título
        worksheet['A1'].font = Font(bold=True, size=14, color="000000")  # Negrito
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')  # Centralizar
        worksheet['A1'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Fundo amarelo

        # Inserir uma linha em branco abaixo do título
        worksheet.insert_rows(2)  # Linha em branco para separar o título dos dados

        # Ajustar a largura das colunas
        for col_letter in worksheet.iter_cols(min_row=3, max_row=3, min_col=1, max_col=len(df.columns)):
            worksheet.column_dimensions[col_letter[0].column_letter].width = 20

        # Formatação do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for cell in worksheet["3:3"]:  # Cabeçalho na linha 3 agora
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill

        # Centralizar o conteúdo das células e adicionar bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Configurar a resposta HTTP com o conteúdo do buffer
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Bloco_{bloco_id}.xlsx"'

    return response

def gerar_excel_todos_blocos(request):
    # Obter todos os blocos
    blocos = Bloco.objects.all()

    # Criar um buffer de memória
    buffer = BytesIO()

    # Gerar o arquivo Excel no buffer
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        for bloco in blocos:
            # Filtrar os apartamentos e residentes correspondentes ao bloco atual
            apartamentos = Apartamento.objects.filter(bloco=bloco)
            residentes = Residente.objects.filter(apartamento__in=apartamentos)

            data = []
            # Iterar sobre os apartamentos e residentes correspondentes
            for apartamento in apartamentos:
                for residente in residentes:
                    if residente.apartamento == apartamento:
                        total_boleto = (residente.valor_aluguel or 0) + (residente.valor_condominio or 0) + (residente.valor_gas or 0) + (residente.outros or 0)

                        # Organizar as informações na ordem desejada
                        entry = {
                            'Apartamento': apartamento.numero,
                            'Residente': residente.nome,
                            'CPF/CNPJ': residente.cpf_cnpj if request.user.is_superuser else None,
                            'Email': residente.email,
                            'Telefone': residente.telefone,
                            'Data Início Contrato': format_date(residente.data_inicio, format="short", locale="pt_BR") if residente.data_inicio else "",
                            'Data Fim Contrato': format_date(residente.data_fim, format="short", locale="pt_BR") if residente.data_fim else "",
                            'Prorrogação': residente.prorrogacao,
                            'Número Cadastro': residente.numero_cadastro,
                            'Valor Aluguel': residente.valor_aluguel,
                            'Valor Condomínio': residente.valor_condominio,
                            'Outros Valores': residente.outros,
                            'Valor Gás': residente.valor_gas,
                            'Data Vencimento Boleto': format_date(residente.data_vencimento, format="short", locale="pt_BR") if residente.data_vencimento else "",
                            'Unidade Consumidora': residente.unidade_consumidora,
                            'Total Boleto': total_boleto,
                        }
                        data.append(entry)

            # Criar o DataFrame para o bloco atual
            if data:  # Verificar se há dados a serem inseridos
                df = pd.DataFrame(data)

                # Adicionar os dados à planilha
                df.to_excel(writer, index=False, sheet_name=f'Bloco {bloco.numero}')

                # Obter a aba ativa (sheet) para formatar
                worksheet = writer.sheets[f'Bloco {bloco.numero}']

                # Adicionar título
                worksheet.insert_rows(1)  # Inserir uma nova linha no topo
                worksheet['A1'] = f'Bloco {bloco.numero}'
                worksheet['A1'].font = Font(bold=True, size=14, color="000000")  # Negrito
                worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')  # Centralizar texto
                worksheet['A1'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Fundo amarelo

                # Inserir uma linha em branco entre o título e o cabeçalho
                worksheet.insert_rows(2)

                # Ajustar a largura das colunas
                for col_letter in worksheet.iter_cols(min_row=3, max_row=3, min_col=1, max_col=len(df.columns)):
                    worksheet.column_dimensions[col_letter[0].column_letter].width = 20

                # Formatação do cabeçalho
                header_font = Font(bold=True, color="FFFFFF")
                header_alignment = Alignment(horizontal='center', vertical='center')  # Centralizar cabeçalhos
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

                for cell in worksheet["3:3"]:  # Cabeçalho
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = header_fill

                # Bordas para as células
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Aplicar bordas e centralizar o conteúdo
                for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Obter data atual para nomear o arquivo
    now = datetime.datetime.now()
    nome_mes = format_datetime(now, "MMMM", locale='pt_BR')  # Nome do mês em português
    ano = now.year  # Ano atual

    # Configurar a resposta HTTP com o conteúdo do buffer
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Planilhas - Blocos | {nome_mes} de {ano}.xlsx"'

    return response
